import locale
import logging
from collections import defaultdict
from openpyxl import load_workbook
import re


def generate_mxtsessions(input_data=None, username="root", encoding="GBK",
                         root_folder="项目名", excel_path=None, sheet_name=None):  # 修改sheet_name默认值为None
    """
    生成MobaXterm会话文件

    参数:
    input_data -- 手动输入的数据列表(可选)
    username -- SSH用户名
    encoding -- 输出文件编码
    root_folder -- 根文件夹名称
    excel_path -- Excel文件路径
    sheet_name -- Excel工作表名称，默认为第一个工作表
    """
    # 设置本地化环境和日志
    locale.setlocale(locale.LC_ALL, 'zh_CN.UTF-8')
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

    output_file = "MobaXterm_Sessions.mxtsessions"
    grouped_data = defaultdict(list)

    # 1. 数据来源处理：优先使用Excel文件
    if excel_path:
        try:
            wb = load_workbook(excel_path)

            # 默认读取第一个工作表
            if sheet_name is None:  # 检查是否未指定工作表名称
                ws = wb.active  # 获取活动工作表（通常是第一个）
                sheet_name = ws.title  # 记录实际使用的工作表名称
                logging.info(f"默认读取第一个工作表: {sheet_name}")
            else:
                ws = wb[sheet_name]  # 使用用户指定的工作表

            for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过标题行
                if not row or not any(row):  # 跳过空行
                    continue

                # 提取前4列有效数据
                parts = [str(cell).strip() if cell is not None else "" for cell in row[:4]]
                # 确保至少有一个非空字段
                if any(parts):
                    grouped_data[parts[0]].append(parts[1:])

            logging.info(f"读取Excel文件成功: {excel_path}, 工作表: {sheet_name}")

        except Exception as e:
            logging.error(f"读取Excel文件失败: {e}")
            return
    elif input_data:
        # 2. 处理手动输入的数据
        for line in input_data:
            if not line.strip():
                continue

            # 更健壮的分割处理
            parts = re.split(r'[,\t]+', line.strip())
            if len(parts) < 4:
                logging.warning(f"忽略无效行: {line}")
                continue

            grouped_data[parts[0]].append(parts[1:4])
    else:
        logging.error("未提供数据来源 (input_data 或 excel_path)")
        return

    # 3. 文件生成
    try:
        with open(output_file, 'w', encoding=encoding) as f:
            # 写入根目录
            root = root_folder.replace('\\', '\\\\') if root_folder else ""
            if root_folder:
                f.write(f"[Bookmarks]\nSubRep={root}\nImgNum=41\n\n")

            # 为每个子路径生成区块
            for idx, (subfolder, devices) in enumerate(grouped_data.items(), start=1):
                if not subfolder:
                    logging.warning("发现空子路径，已跳过")
                    continue

                # 路径安全处理
                safe_subfolder = re.sub(r'[\\/*?:"<>|]', "_", subfolder)
                safe_subfolder = safe_subfolder.replace("（", "(").replace("）", ")")

                # 构建路径层级
                path_segments = [seg for seg in [root_folder, safe_subfolder] if seg]
                full_path = "\\\\".join(path_segments)

                f.write(f"[Bookmarks_{idx}]\n")
                f.write(f"SubRep={full_path}\n")
                f.write("ImgNum=41\n")

                for device in devices:
                    # 确保最多取3个值
                    ip, device_id, ssh_addr = device[:3] if len(device) >= 3 else (device[0], "", "")

                    # 解析SSH地址
                    if ssh_addr and ':' in ssh_addr:
                        ssh_host, ssh_port = ssh_addr.split(':', 1)
                        logging.info(f"使用穿透IP连接: {ssh_host}，端口{ssh_port}")
                    else:
                        logging.info(f"使用远程IP连接: {ip}，端口22")
                        ssh_host = ssh_addr if ssh_addr else ip  # 使用IP作为主机名
                        ssh_port = "22"

                    # 安全处理会话名称
                    session_name = f"{ip}-{device_id}".replace(' ', '_') if device_id else ip

                    # 写入会话配置
                    config_line = (
                        f"{session_name}=#109#0%{ssh_host}%{ssh_port}%{username}%%-1%-1%%%%%0%0%0%%%-1%0%0%0%%1080%%0%0%1"
                        f"#MobaFont%10%0%0%-1%15%236,236,236%30,30,30%180,180,192%0%-1%0%%xterm%-1%0%"
                        f"_Std_Colors_0_%80%24%0%1%-1%<none>%%0%0%-1%-1#0# #-1\n"
                    )
                    f.write(config_line)
                f.write("\n")  # 组间空行分隔

        logging.info(f"文件已生成: {output_file}，请使用MobaXterm导入")

    except Exception as e:
        logging.error(f"生成文件失败: {e}")


def main():
    """交互式生成 MobaXterm 会话配置文件"""
    # 1. 提示并选择文件（支持默认值和路径检查）
    default_excel = "import.xlsx"
    while True:
        excel_path = input(f"请输入Excel文件路径（默认：{default_excel}）：").strip() or default_excel
        if not excel_path.endswith('.xlsx'):
            excel_path += '.xlsx'
        try:
            open(excel_path, 'rb').close()  # 简单验证文件可访问性
            break
        except FileNotFoundError:
            print(f"错误：文件 {excel_path} 不存在，请重新输入！")
        except Exception as e:
            print(f"文件访问错误：{str(e)}")

    # 2. 提示输入SSH用户名（带默认值）
    default_user = "root"
    username = input(f"请输入SSH连接用户名（默认：{default_user}）：").strip() or default_user

    default_folder = "项目名称"
    # 3. 提示输入项目名称（支持空值）
    root_folder = input("请输入项目名称（直接回车可跳过）：").strip() or default_folder

    # 4. 执行生成操作
    try:
        print("\n正在生成会话文件，请稍候...")
        generate_mxtsessions(
            excel_path=excel_path,
            username=username,
            root_folder=root_folder
        )
        print("生成完成！请打开 MobaXterm 导入文件")
        input("\n按回车键退出程序...")
    except Exception as e:
        print(f"生成过程中出错：{str(e)}")
        logging.error(f"生成失败: {e}")

if __name__ == '__main__':
    main()